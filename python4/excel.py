import csv

from openpyxl import load_workbook
from openpyxl import Workbook
import json,pandas

excel = load_workbook('C:\\Users\\linzh\\Desktop\\编剧素材\成功学\\工作.xlsx')
sheet = excel["Sheet1"]
lst=[]

for cell in list(sheet.rows)[9]:
    lst.append(cell.value)
print(lst)

with open("./hmm.json",'w',encoding='utf-8') as json_file:
    json.dump(lst,json_file,ensure_ascii=False)

with open('example.csv', 'w', newline='') as f:
    writer = csv.writer(f)
    for row in lst:
        writer.writerow(row)

wb = Workbook()
wb.remove(wb.active)
ws1 = wb.create_sheet("mysheet",index=1)
ws1.append(lst)
wb.save('C:\\Users\\linzh\\Desktop\\编剧素材\成功学\\pythonExcel.xlsx')
